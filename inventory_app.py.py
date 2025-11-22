import customtkinter as ctk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

# --- Config ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class InventoryApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Window Setup
        self.title("E-WerHaus")
        self.geometry("800x680") 
        self.nama_file_db = "data_inventory.xlsx"
        
        self.cek_database()

        # UI Layout
        self.tabview = ctk.CTkTabview(self, width=750, height=600)
        self.tabview.pack(padx=20, pady=(20, 5))

        self.tab_input = self.tabview.add("Input Barang")
        self.tab_data = self.tabview.add("Data Gudang")

        # Copyright Footer
        self.label_copyright = ctk.CTkLabel(self, text="Copyright © Albani Computer", font=("Arial", 11), text_color="gray")
        self.label_copyright.pack(side="bottom", pady=10)

        self.setup_tab_input()
        self.setup_tab_data()
        self.muat_data()

    def setup_tab_input(self):
        # Header
        ctk.CTkLabel(self.tab_input, text="Tambah Stok Baru", font=("Roboto", 20, "bold")).pack(pady=20)

        # Form Inputs
        self.entry_nama = ctk.CTkEntry(self.tab_input, placeholder_text="Nama Barang", width=300)
        self.entry_nama.pack(pady=10)

        self.entry_jumlah = ctk.CTkEntry(self.tab_input, placeholder_text="Jumlah (Stok)", width=300)
        self.entry_jumlah.pack(pady=10)

        self.entry_harga = ctk.CTkEntry(self.tab_input, placeholder_text="Harga Satuan (Rp)", width=300)
        self.entry_harga.pack(pady=10)

        # Action Buttons
        ctk.CTkButton(self.tab_input, text="Simpan Data", command=self.simpan_data, width=300).pack(pady=20)
        
        self.label_status = ctk.CTkLabel(self.tab_input, text="", text_color="green")
        self.label_status.pack(pady=5)

    def setup_tab_data(self):
        # Search Bar
        frame_cari = ctk.CTkFrame(self.tab_data, fg_color="transparent")
        frame_cari.pack(pady=10, fill="x")

        self.entry_cari = ctk.CTkEntry(frame_cari, placeholder_text="Cari nama barang...", width=200)
        self.entry_cari.pack(side="left", padx=(0, 10), expand=True, fill="x")

        ctk.CTkButton(frame_cari, text="Cari", command=self.lakukan_pencarian, width=80).pack(side="left", padx=(0, 10))
        ctk.CTkButton(frame_cari, text="Reset", command=self.muat_data, width=80, fg_color="#555555").pack(side="left")

        # Treeview Style
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background="#2b2b2b", foreground="white", fieldbackground="#2b2b2b", rowheight=25)
        style.map('Treeview', background=[('selected', '#1f538d')])
        
        # Table Setup
        columns = ("nama", "jumlah", "harga")
        self.tabel = ttk.Treeview(self.tab_data, columns=columns, show="headings", height=15)
        
        self.tabel.heading("nama", text="Nama Barang")
        self.tabel.heading("jumlah", text="Jumlah")
        self.tabel.heading("harga", text="Harga")

        self.tabel.column("nama", width=300)
        self.tabel.column("jumlah", width=100, anchor="center")
        self.tabel.column("harga", width=150, anchor="e")

        self.tabel.pack(pady=10, padx=10, fill="both", expand=True)

        # Action Buttons (Edit/Delete)
        frame_aksi = ctk.CTkFrame(self.tab_data, fg_color="transparent")
        frame_aksi.pack(pady=10)

        ctk.CTkButton(frame_aksi, text="Edit Barang", command=self.buka_jendela_edit, fg_color="#D35400", hover_color="#A04000", width=150).pack(side="left", padx=10)
        ctk.CTkButton(frame_aksi, text="Hapus Barang", command=self.hapus_data, fg_color="#C0392B", hover_color="#922B21", width=150).pack(side="left", padx=10)

    # --- Database Logic ---

    def cek_database(self):
        if not os.path.exists(self.nama_file_db):
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Barang"
            ws.append(["Nama Barang", "Jumlah", "Harga"])
            wb.save(self.nama_file_db)

    def simpan_data(self):
        nama = self.entry_nama.get()
        jumlah = self.entry_jumlah.get()
        harga = self.entry_harga.get()

        if not all([nama, jumlah, harga]):
            self.label_status.configure(text="Error: Isi semua data!", text_color="red")
            return

        try:
            wb = load_workbook(self.nama_file_db)
            ws = wb.active
            ws.append([nama, jumlah, harga])
            wb.save(self.nama_file_db)

            self.label_status.configure(text=f"Berhasil: {nama} disimpan!", text_color="#2CC985")
            self.clear_inputs()
            self.muat_data() 

        except Exception as e:
            self.label_status.configure(text=f"Error: {str(e)}", text_color="red")

    def muat_data(self, keyword=""):
        for item in self.tabel.get_children():
            self.tabel.delete(item)

        try:
            wb = load_workbook(self.nama_file_db)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                nama_cell = str(row[0]).lower() if row[0] else ""
                search_key = keyword.lower()
                
                if search_key == "" or search_key in nama_cell:
                    self.tabel.insert("", "end", values=row)
        
        except Exception as e:
            print(f"Error loading data: {e}")

    def lakukan_pencarian(self):
        self.muat_data(self.entry_cari.get())

    def hapus_data(self):
        selected = self.tabel.selection()
        if not selected:
            messagebox.showwarning("Peringatan", "Pilih barang yang ingin dihapus!")
            return

        nama_barang = self.tabel.item(selected)['values'][0]
        
        if messagebox.askyesno("Konfirmasi", f"Yakin ingin menghapus '{nama_barang}'?"):
            try:
                wb = load_workbook(self.nama_file_db)
                ws = wb.active
                
                for row in ws.iter_rows(min_row=2):
                    if str(row[0].value) == str(nama_barang):
                        ws.delete_rows(row[0].row)
                        wb.save(self.nama_file_db)
                        messagebox.showinfo("Sukses", "Data berhasil dihapus!")
                        self.muat_data()
                        return
            except Exception as e:
                messagebox.showerror("Error", f"Gagal menghapus: {e}")

    def buka_jendela_edit(self):
        selected = self.tabel.selection()
        if not selected:
            messagebox.showwarning("Peringatan", "Pilih barang yang ingin diedit!")
            return

        values = self.tabel.item(selected)['values']
        self.create_edit_window(values)

    def create_edit_window(self, values):
        self.window_edit = ctk.CTkToplevel(self)
        self.window_edit.title("Edit Barang")
        self.window_edit.geometry("400x400")
        self.window_edit.attributes("-topmost", True) 

        ctk.CTkLabel(self.window_edit, text="Edit Data Barang", font=("Roboto", 18, "bold")).pack(pady=20)

        self.entry_edit_nama = ctk.CTkEntry(self.window_edit, width=250)
        self.entry_edit_nama.insert(0, values[0])
        self.entry_edit_nama.pack(pady=10)

        self.entry_edit_jumlah = ctk.CTkEntry(self.window_edit, width=250)
        self.entry_edit_jumlah.insert(0, values[1])
        self.entry_edit_jumlah.pack(pady=10)

        self.entry_edit_harga = ctk.CTkEntry(self.window_edit, width=250)
        self.entry_edit_harga.insert(0, values[2])
        self.entry_edit_harga.pack(pady=10)

        ctk.CTkButton(self.window_edit, text="Simpan Perubahan", 
                      command=lambda: self.proses_update_data(values[0])).pack(pady=20)

    def proses_update_data(self, nama_asli):
        nama_baru = self.entry_edit_nama.get()
        jumlah_baru = self.entry_edit_jumlah.get()
        harga_baru = self.entry_edit_harga.get()

        try:
            wb = load_workbook(self.nama_file_db)
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == str(nama_asli):
                    row[0].value = nama_baru
                    row[1].value = jumlah_baru
                    row[2].value = harga_baru
                    wb.save(self.nama_file_db)
                    
                    messagebox.showinfo("Sukses", "Data berhasil diupdate!")
                    self.window_edit.destroy()
                    self.muat_data()
                    return
            
            messagebox.showerror("Error", "Data asli tidak ditemukan.")

        except Exception as e:
            messagebox.showerror("Error", f"Gagal update: {e}")

    def clear_inputs(self):
        self.entry_nama.delete(0, 'end')
        self.entry_jumlah.delete(0, 'end')
        self.entry_harga.delete(0, 'end')

if __name__ == "__main__":
    app = InventoryApp()
    app.mainloop()